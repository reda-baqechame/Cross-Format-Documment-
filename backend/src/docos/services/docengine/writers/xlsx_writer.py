"""Universal canonical-model → XLSX writer.

Rebuilds a real ``.xlsx`` from the node graph. Each :class:`TableNode` becomes a
worksheet; a heading immediately preceding a table names its sheet. Free-flowing
text blocks (paragraphs / headings with no table) are collected onto a final
"Text" sheet, one row per block, so nothing is silently dropped. Like the DOCX
writer it is deliberately tolerant: unknown nodes are skipped rather than raising.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from docos.model.document import CanonicalDocument
from docos.model.nodes import AnyNode
from docos.services.docengine.writers.redaction import run_text, spreadsheet_text

# openpyxl rejects sheet titles over 31 chars or containing []:*?/\
_BAD_TITLE = set('[]:*?/\\')


def _block_text(doc: CanonicalDocument, block: AnyNode) -> str:
    return "".join(run_text(doc, r) for r in doc.children_of(block.id) if r.type == "run")


def _content_nodes(doc: CanonicalDocument) -> list[AnyNode]:
    """Top-level content blocks, descending through ``page`` containers.

    PDF and PPTX adapters nest their content under ``page`` nodes; iterating only the root's
    direct children would miss all of it and produce an empty workbook. This flattens one page
    level (mirroring the DOCX writer) so paged documents export their actual content.
    """
    out: list[AnyNode] = []
    for node in doc.children_of(doc.root_id):
        if node.type == "page":
            out.extend(doc.children_of(node.id))
        else:
            out.append(node)
    return out


def _safe_title(name: str, used: set[str]) -> str:
    cleaned = "".join(" " if ch in _BAD_TITLE else ch for ch in name).strip() or "Sheet"
    cleaned = cleaned[:31]
    base, n = cleaned, 2
    while cleaned.lower() in used:
        suffix = f" ({n})"
        cleaned = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned.lower())
    return cleaned


def _write_table(ws: Worksheet, doc: CanonicalDocument, tnode: AnyNode) -> None:
    rows = [n for n in doc.children_of(tnode.id) if n.type == "table_row"]
    for ri, row in enumerate(rows, start=1):
        cells = [c for c in doc.children_of(row.id) if c.type == "table_cell"]
        for ci, cell in enumerate(cells, start=1):
            # A cell may carry a formula string (e.g. "=A1+B1"); write it as a real Excel formula
            # so Excel recomputes it on open. We don't recompute in-app (no calc engine), so the
            # displayed text stays the last-known value until Excel recalculates.
            formula = cell.attrs.get("formula")
            if isinstance(formula, str) and formula.startswith("="):
                ws.cell(row=ri, column=ci, value=formula)
            else:
                ws.cell(row=ri, column=ci, value=spreadsheet_text(_block_text(doc, cell)))
            fmt = cell.attrs.get("number_format")
            if fmt:
                ws.cell(row=ri, column=ci).number_format = fmt


def model_to_xlsx(doc: CanonicalDocument) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet; we add our own
    if doc.meta.title:
        wb.properties.title = doc.meta.title

    used_titles: set[str] = set()
    pending_heading: str | None = None
    loose_text: list[str] = []
    made_sheet = False

    for node in _content_nodes(doc):
        kind = node.type
        if kind == "heading":
            pending_heading = _block_text(doc, node) or None
        elif kind == "table":
            title = _safe_title(pending_heading or "Sheet", used_titles)
            _write_table(wb.create_sheet(title=title), doc, node)
            pending_heading = None
            made_sheet = True
        elif kind in ("paragraph", "list", "list_item"):
            if pending_heading:
                loose_text.append(pending_heading)
                pending_heading = None
            if kind == "list":
                for item in doc.children_of(node.id):
                    if item.type == "list_item":
                        loose_text.append(_block_text(doc, item))
            else:
                text = _block_text(doc, node)
                if text:
                    loose_text.append(text)

    if pending_heading:
        loose_text.append(pending_heading)
    if loose_text:
        ws = wb.create_sheet(title=_safe_title("Text", used_titles))
        for i, line in enumerate(loose_text, start=1):
            ws.cell(row=i, column=1, value=spreadsheet_text(line))
        made_sheet = True
    if not made_sheet:
        wb.create_sheet(title="Sheet1")  # never emit a zero-sheet (invalid) workbook

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
