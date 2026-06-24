"""XLSX adapter (openpyxl).

Maps each worksheet to a heading (the sheet name) followed by a table built from the
sheet's used range; every cell becomes a table cell carrying its displayed value as a
run. Number formats are preserved in ``attrs`` so nothing is silently dropped. Once in
the canonical model, a spreadsheet exports to DOCX/TXT like any other document.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

from openpyxl import load_workbook

from docos.model.document import CanonicalDocument, DocumentMeta
from docos.model.ids import new_doc_id, new_node_id
from docos.model.nodes import (
    HeadingNode,
    RootNode,
    RunNode,
    TableCellNode,
    TableNode,
    TableRowNode,
)
from docos.services.docengine.interface import FormatAdapter
from docos.storage.blob import BlobStore

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# Bound very large sheets so a single upload can't blow up the model.
_MAX_ROWS = 500
_MAX_COLS = 50


class XlsxAdapter(FormatAdapter):
    format_id = "xlsx"
    supported_mimes = (_XLSX_MIME,)

    def parse(self, data: bytes, *, blob: BlobStore | None = None) -> CanonicalDocument:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        now = datetime.now(UTC)
        root = RootNode(id=new_node_id("root"))
        props = wb.properties
        meta = DocumentMeta(
            title=props.title or None,
            author=props.creator or None,
            source_format="xlsx",
            source_mime=_XLSX_MIME,
            created_at=now,
            modified_at=now,
            page_count=len(wb.sheetnames),
            custom={"last_modified_by": props.lastModifiedBy, "keywords": props.keywords},
        )
        doc = CanonicalDocument(doc_id=new_doc_id(), root_id=root.id, meta=meta)
        doc.add_node(root)

        order = 0
        for sheet in wb.worksheets:
            heading = HeadingNode(id=new_node_id(), parent_id=root.id, level=2, reading_order=order)
            heading.tags.append("H2")
            htext = RunNode(id=new_node_id(), parent_id=heading.id, text=sheet.title, bold=True)
            heading.children.append(htext.id)
            root.children.append(heading.id)
            doc.add_node(heading)
            doc.add_node(htext)
            order += 1
            order = self._parse_sheet(doc, root, sheet, order)

        if meta.title:
            doc.accessibility.has_doc_title = True
        return doc

    def _parse_sheet(self, doc: CanonicalDocument, root, sheet, order: int) -> int:
        nrows = min(sheet.max_row or 0, _MAX_ROWS)
        ncols = min(sheet.max_column or 0, _MAX_COLS)
        if nrows == 0 or ncols == 0:
            return order

        tnode = TableNode(
            id=new_node_id(), parent_id=root.id, rows=nrows, cols=ncols, reading_order=order
        )
        for r, row in enumerate(sheet.iter_rows(max_row=nrows, max_col=ncols)):
            rownode = TableRowNode(id=new_node_id(), parent_id=tnode.id, row=r)
            for c, cell in enumerate(row):
                cellnode = TableCellNode(id=new_node_id(), parent_id=rownode.id, row=r, col=c)
                value = "" if cell.value is None else str(cell.value)
                run = RunNode(id=new_node_id(), parent_id=cellnode.id, text=value)
                if cell.number_format and cell.number_format != "General":
                    cellnode.attrs["number_format"] = cell.number_format
                cellnode.children.append(run.id)
                rownode.children.append(cellnode.id)
                doc.add_node(cellnode)
                doc.add_node(run)
            tnode.children.append(rownode.id)
            doc.add_node(rownode)
        root.children.append(tnode.id)
        doc.add_node(tnode)
        return order + 1

    def render_preview(self, doc: CanonicalDocument, page: int) -> bytes:
        raise NotImplementedError("XlsxAdapter.render_preview — sheets render in the canvas")

    def export(self, doc: CanonicalDocument, *, target_mime: str) -> bytes:
        from docos.services.docengine.writers.xlsx_writer import model_to_xlsx

        return model_to_xlsx(doc)
