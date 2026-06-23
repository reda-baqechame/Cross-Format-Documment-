"""PDF table extraction: a ruled grid becomes TableNodes and exports to a real xlsx sheet."""

from __future__ import annotations

import io

import fitz
from openpyxl import load_workbook

from docos.services.docengine.adapters.pdf import PdfAdapter
from docos.services.docengine.writers.xlsx_writer import model_to_xlsx


def _grid_pdf() -> bytes:
    """A 2x2 grid with ruled lines + cell text — what find_tables() is designed to detect."""
    pdf = fitz.open()
    page = pdf.new_page(width=400, height=300)
    # Cell text.
    cells = {(0, 0): "Item", (1, 0): "Qty", (0, 1): "Widget", (1, 1): "5"}
    x0, y0, cw, ch = 50, 50, 120, 40
    for (col, row), text in cells.items():
        page.insert_text((x0 + col * cw + 5, y0 + row * ch + 25), text, fontsize=12)
    # Ruled lines (grid) so the table detector locks on.
    for r in range(3):
        page.draw_line((x0, y0 + r * ch), (x0 + 2 * cw, y0 + r * ch))
    for c in range(3):
        page.draw_line((x0 + c * cw, y0), (x0 + c * cw, y0 + 2 * ch))
    data = pdf.tobytes()
    pdf.close()
    return data


def test_pdf_table_is_parsed_into_table_nodes():
    doc = PdfAdapter().parse(_grid_pdf())
    tables = [n for n in doc.nodes.values() if n.type == "table"]
    assert len(tables) == 1
    rows = [n for n in doc.nodes.values() if n.type == "table_row"]
    cells = [n for n in doc.nodes.values() if n.type == "table_cell"]
    assert len(rows) == 2
    assert len(cells) == 4


def test_pdf_table_exports_to_xlsx_with_values():
    doc = PdfAdapter().parse(_grid_pdf())
    wb = load_workbook(io.BytesIO(model_to_xlsx(doc)))
    # Some sheet must carry the cell values (not the empty fallback).
    all_values = {
        str(cell.value).strip()
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert {"Item", "Qty", "Widget", "5"} <= all_values
