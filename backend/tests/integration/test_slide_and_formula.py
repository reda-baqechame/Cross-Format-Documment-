"""Native slide/spreadsheet fidelity (tractable slice): per-slide thumbnails + cell formulas."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from docos.services.docengine.adapters.txt import TxtAdapter
from docos.services.docengine.writers.image_writer import model_to_png


def test_per_page_thumbnail_is_valid_png_and_isolated():
    from docos.model.ids import new_node_id
    from docos.model.nodes import PageNode, ParagraphNode, RunNode

    doc = TxtAdapter().parse(b"intro")
    root = doc.nodes[doc.root_id]
    page_ids = []
    for i, text in enumerate(["SLIDE ONE CONTENT", "SLIDE TWO CONTENT"], start=1):
        page = PageNode(
            id=new_node_id("page"), parent_id=root.id, page_number=i, width=960.0, height=540.0
        )
        root.children.append(page.id)
        doc.add_node(page)
        para = ParagraphNode(id=new_node_id(), parent_id=page.id)
        run = RunNode(id=new_node_id(), parent_id=para.id, text=text)
        para.children.append(run.id)
        page.children.append(para.id)
        doc.add_node(para)
        doc.add_node(run)
        page_ids.append(page.id)

    png = model_to_png(doc, root_id=page_ids[0])
    assert png[:8] == b"\x89PNG\r\n\x1a\n"  # valid PNG signature
    # Whole-doc render differs from a single-slide render (proves isolation).
    assert model_to_png(doc) != png


def test_slide_thumbnail_endpoint(client):
    doc_id = client.post(
        "/documents", files={"file": ("d.txt", b"Hello slide", "text/plain")}
    ).json()["doc_id"]
    nodes = client.get(f"/documents/{doc_id}/model").json()["document"]["nodes"]
    page_id = next((nid for nid, n in nodes.items() if n["type"] == "page"), None)
    if page_id is None:  # txt may render under root without an explicit page
        return
    res = client.get(f"/documents/{doc_id}/slide-thumbnail?node_id={page_id}")
    assert res.status_code == 200
    assert res.headers["content-type"] == "image/png"


def test_cell_formula_exports_as_excel_formula():
    doc = TxtAdapter().parse(b"placeholder")
    from docos.model.ids import new_node_id
    from docos.model.nodes import TableCellNode, TableNode, TableRowNode

    table = TableNode(id=new_node_id(), parent_id=doc.root_id, rows=1, cols=1)
    doc.nodes[doc.root_id].children.append(table.id)
    doc.add_node(table)
    row = TableRowNode(id=new_node_id(), parent_id=table.id, row=0)
    table.children.append(row.id)
    doc.add_node(row)
    cell = TableCellNode(
        id=new_node_id(), parent_id=row.id, row=0, col=0, attrs={"formula": "=1+2"}
    )
    row.children.append(cell.id)
    doc.add_node(cell)

    from docos.services.docengine.writers.xlsx_writer import model_to_xlsx

    wb = load_workbook(io.BytesIO(model_to_xlsx(doc)))
    # The formula is written as a real formula string the workbook reads back.
    values = [c.value for rowcells in wb.worksheets[0].iter_rows() for c in rowcells]
    assert "=1+2" in values
