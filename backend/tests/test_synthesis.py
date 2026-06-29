"""Document synthesis: pack findings → a generated, writer-ready document in every format."""

from __future__ import annotations

import io

from docos.services import synthesis
from docos.services.docengine.writers.markup import model_to_markdown
from docos.services.docengine.writers.searchable_pdf import model_to_searchable_pdf
from docos.services.docengine.writers.xlsx_writer import model_to_xlsx
from docos.services.packs.finance import APMatch, APReport
from docos.services.packs.import_export import PacketFinding, ShipmentFields


def _ap_report() -> APReport:
    return APReport(
        document_count=2,
        documents=[
            ShipmentFields(
                doc_id="inv1",
                title="Invoice",
                doc_type="invoice",
                confidence=0.6,
                currency="USD",
                total=100.0,
                invoice_number="INV-42",
                po_number="PO-9",
            ),
            ShipmentFields(
                doc_id="po1", title="PO", doc_type="purchase_order", confidence=0.6, total=120.0
            ),
        ],
        matches=[
            APMatch(
                invoice_doc_id="inv1",
                po_number="PO-9",
                matched_po_doc_id="po1",
                total_matches=False,
                currency_matches=None,
            )
        ],
        findings=[
            PacketFinding(severity="error", code="po_total_mismatch", message="100 ≠ 120 on PO-9.")
        ],
        summary="1 blocking AP issue(s); 1/1 invoice(s) matched to a PO.",
    )


def test_build_document_is_walkable_and_titled():
    doc = synthesis.build_document(synthesis.ap_reconciliation_report(_ap_report()))
    assert doc.meta.title == "Accounts-Payable Reconciliation"
    # Root has ordered children and the graph is internally consistent (every child resolves).
    assert doc.nodes[doc.root_id].children
    for node in doc.nodes.values():
        for child in node.children:
            assert child in doc.nodes


def test_report_renders_to_markdown_with_findings():
    doc = synthesis.build_document(synthesis.ap_reconciliation_report(_ap_report()))
    md = model_to_markdown(doc).decode()
    assert "Accounts-Payable Reconciliation" in md
    assert "INV-42" in md  # the extracted field flows into the deliverable
    assert "po_total_mismatch" in md  # the finding is rendered


def test_report_renders_to_xlsx_and_pdf():
    doc = synthesis.build_document(synthesis.ap_reconciliation_report(_ap_report()))

    xlsx = model_to_xlsx(doc)
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx))
    flat = " ".join(
        str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value
    )
    assert "INV-42" in flat or "po_total_mismatch" in flat

    pdf = model_to_searchable_pdf(doc)
    assert pdf[:4] == b"%PDF"


def test_generated_report_never_leaks_redacted_text():
    # Redaction is true removal: a redacted line must not reach pack findings, so it can't reach the
    # generated deliverable. Build a real doc, redact the secret line, run the pack, synthesize.
    from datetime import UTC, datetime

    from docos.model.document import CanonicalDocument, DocumentMeta
    from docos.model.ids import new_doc_id, new_node_id
    from docos.model.nodes import ParagraphNode, RootNode, RunNode
    from docos.services.packs import check_ap

    now = datetime.now(UTC)
    root = RootNode(id=new_node_id("root"))
    doc = CanonicalDocument(
        doc_id=new_doc_id(),
        root_id=root.id,
        meta=DocumentMeta(
            source_format="txt", source_mime="text/plain", created_at=now, modified_at=now
        ),
    )
    doc.add_node(root)
    secret_id = ""
    for i, line in enumerate(
        ["Commercial Invoice", "Invoice number: SECRET-9", "Total due: 9 USD"]
    ):
        p = ParagraphNode(id=new_node_id(), parent_id=root.id, reading_order=i)
        r = RunNode(id=new_node_id(), parent_id=p.id, text=line)
        p.children.append(r.id)
        root.children.append(p.id)
        doc.add_node(p)
        doc.add_node(r)
        if "SECRET-9" in line:
            secret_id = r.id
    doc.redaction.redacted_node_ids.append(secret_id)

    report = synthesis.ap_reconciliation_report(check_ap([("d1", "inv", doc)]))
    md = model_to_markdown(synthesis.build_document(report)).decode()
    assert "SECRET-9" not in md


def test_packet_report_adapter_smoke():
    from docos.services.packs.import_export import ChecklistItem, PacketReport

    rep = PacketReport(
        document_count=1,
        documents=[
            ShipmentFields(
                doc_id="d1",
                title=None,
                doc_type="commercial_invoice",
                confidence=0.5,
                currency="EUR",
            )
        ],
        findings=[PacketFinding(severity="warn", code="hs_code_missing", message="No HS code.")],
        checklist=[ChecklistItem(doc_type="packing_list", label="Packing list", present=False)],
        summary="No blocking issues; 1 warning(s).",
    )
    md = model_to_markdown(
        synthesis.build_document(synthesis.packet_exception_report(rep))
    ).decode()
    assert "Exception Report" in md and "hs_code_missing" in md and "Packing list" in md
